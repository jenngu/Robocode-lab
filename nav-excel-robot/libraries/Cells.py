# +
import openpyxl as xl

class Cells:
    def copy_cells():
        path1 = C:\\Users\\jenngu\\Robots\\nav-excel-robot\\devdata\\workbook1.xlsx
        path2 = C:\\Users\\jenngu\\Robots\\nav-excel-robot\\tasks\\workbook2.xlsx
        
        wb1 = xl.load_workbook(filename=path1)
        ws1 = wb1.worksheets[0]

        wb2 = xl.load_workbook(filename=path2)
        ws2 = wb2.create_sheet(ws1.title)
        
        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value
        
        wb2.save(path2)

